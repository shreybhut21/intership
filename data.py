# import pandas
from flask import Flask, render_template, request, redirect, url_for, flash
import os
import openpyxl
import csv
import uuid
from datetime import datetime

app = Flask(__name__)
app.secret_key = "your_secret_key"
excel_file = "user_form.xlsx"
csv_file = "user_data.csv"
fields = ["UID", "Name", "Email", "Phone", "Password", "Country", "Hobbies", "Duration", "DOB", "Rating", "City", "State", "Zip Code"]

def read_excel_file(filename=excel_file):
    if not os.path.exists(filename):
        return []
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(row):
            data.append({field: row[i] if i < len(row) else "" for i, field in enumerate(fields)})
    return data

def write_excel_file(users, filename=excel_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(fields)
    for user in users:
        sheet.append([user.get(field, "") for field in fields])
    workbook.save(filename)

def write_csv_file(users, filename=csv_file):
    with open(filename, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for user in users:
            writer.writerow(user)

def find_user_index_by_id(users, user_id):
    for idx, user in enumerate(users):
        if user["UID"] == user_id:
            return idx
    return None

def update_excel_row(row_index, new_row, filename=excel_file):
    try:
        book = openpyxl.load_workbook(filename)
        sheet = book.active
        for col, key in enumerate(fields, start=1):
            sheet.cell(row=row_index, column=col, value=new_row.get(key, ""))
        book.save(filename)
    except Exception as e:
        print(f"Error updating Excel: {e}")

def get_country_from_code(code):
    if code == "+91":
        return "India"
    elif code == "+1":
        return "United States"
    elif code == "+44":
        return "United Kingdom"
    elif code == "+81":
        return "Japan"
    elif code == "+61":
        return "Australia"
    else:
        return "no country"

@app.route("/", methods=["GET", "POST"])
def main():
    errors = {}
    form_data = {}
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        country_code = request.form.get("country_code", "").strip()
        phone = request.form.get("phone", "").strip()
        dob = request.form.get("DOB", "").strip()
        full_phone = country_code + phone  
        form_data = {
            "name": name,
            "email": request.form.get("email", ""),
            "country_code": country_code,
            "phone": phone,
            "city": request.form.get("city", ""),
            "state": request.form.get("state", ""),
            "zip": request.form.get("zip", ""),
            "rating": request.form.get("rating", "5"),
            "hobby": request.form.getlist("hobby"),
            "duration": request.form.get("duration", ""),
            "DOB": dob
        }

        if country_code == "+91":
            dob_limit = "2006-01-01"
        else:
            dob_limit = "1995-01-01"
        try:
            dob_date = datetime.strptime(dob, "%Y-%m-%d").date()
            dob_limit_date = datetime.strptime(dob_limit, "%Y-%m-%d").date()
            if dob_date > dob_limit_date:
                errors["DOB"] = "You are under age."
        except Exception:
            errors["DOB"] = "Invalid date format."
        if not dob:
            errors["DOB"] = "DOB is required."
        if not name or len(name) < 2:
            errors["name"] = "Name is too short."
        if not phone.isdigit() or len(phone) < 10:
            errors["phone"] = "Phone number is incorrect."
        if not dob:
            errors["DOB"] = "DOB is required."
        elif dob > "2005-01-01":
            errors["DOB"] = "You are under age."
        
        if not errors:
            country = get_country_from_code(country_code)
            user = {
                "UID": str(uuid.uuid4()),
                "Name": name,
                "Email": request.form.get("email", ""),
                "Phone": full_phone,
                "Password": request.form.get("password", ""),
                "Country": country,  
                "Hobbies": ", ".join(request.form.getlist("hobby")),
                "Duration": request.form.get("duration", ""),
                "DOB": dob,
                "Rating": request.form.get("rating", ""),
                "City": request.form.get("city", ""),
                "State": request.form.get("state", ""),
                "Zip Code": request.form.get("zip", ""),
            }
            users = read_excel_file()
            users.append(user)
            users.sort(key=lambda x: (x["Name"] or "").lower())
            write_excel_file(users)
            write_csv_file(users)
            return redirect(url_for("table"))
    else:
        form_data = {}
    users = read_excel_file()
    return render_template("main.html", users=users, errors=errors, form_data=form_data)

@app.route("/table")
def table():
    users = read_excel_file()
    users.sort(key=lambda x: (x["Name"] or "").lower())
    return render_template("table.html", users=users)


@app.route("/edit/<user_id>", methods=["GET", "POST"])
def edit_user(user_id):
    users = read_excel_file()
    index = find_user_index_by_id(users, user_id)
    if index is None:
        return "no user"
    user = users[index]
    errors = {}
    phone = user.get("Phone", "")
    country_code = phone[:3] if phone.startswith("+") else ""
    phone_number = phone[3:] if phone.startswith("+") else phone

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        country_code = request.form.get("country_code", "").strip()
        phone = request.form.get("phone", "").strip()
        dob = request.form.get("DOB", "").strip()
        full_phone = country_code + phone

        if not name or len(name) < 2:
            errors["name"] = "Name is too short."
        if not phone.isdigit() or len(phone) < 10:
            errors["phone"] = "Phone number is incorrect."
        if not dob:
            errors["DOB"] = "DOB is required."
        elif dob > "2005-01-01":
            errors["DOB"] = "You are under age."

        if not errors:
            country = get_country_from_code(country_code)
            user["Name"] = name
            user["Email"] = request.form.get("email", "")
            user["Phone"] = full_phone
            user["Password"] = request.form.get("password", "")
            user["Country"] = country
            user["Hobbies"] = ", ".join(request.form.getlist("hobby"))
            user["Duration"] = request.form.get("duration", "")
            user["DOB"] = dob
            user["Rating"] = request.form.get("rating", "")
            user["City"] = request.form.get("city", "")
            user["State"] = request.form.get("state", "")
            user["Zip Code"] = request.form.get("zip", "")
            users[index] = user
            write_csv_file(users)
            update_excel_row(index + 2, user)
            return redirect(url_for("table"))

    hobbies = []
    if "Hobbies" in user and isinstance(user["Hobbies"], str):
        hobbies = user["Hobbies"].split(", ")
    user_for_form = {
        "name": user.get("Name", ""),
        "email": user.get("Email", ""),
        "country_code": country_code,
        "phone": phone_number,
        "password": user.get("Password", ""),
        "city": user.get("City", ""),
        "state": user.get("State", ""),
        "zip": user.get("Zip Code", ""),
        "rating": user.get("Rating", ""),
        "hobby": hobbies,
        "duration": user.get("Duration", ""),
        "DOB": user.get("DOB", "")
    }
    return render_template("edit.html", user=user_for_form, index=index, errors=errors)

@app.route("/delete/<user_id>")
def delete_user(user_id):
    users = read_excel_file()
    index = find_user_index_by_id(users, user_id)
    if index is None:
        return "User not found", 404
    users.pop(index)
    if users:
        users.sort(key=lambda x: (x["Name"] or "").lower())
        write_csv_file(users)
        write_excel_file(users)
    else:
        if os.path.exists(excel_file):
            os.remove(excel_file)
        if os.path.exists(csv_file):
            os.remove(csv_file)
    return redirect(url_for("table"))


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)